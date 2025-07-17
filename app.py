from flask import Flask, render_template, request, send_file
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from kontroller import zorunluluk_dict,tamsayi_alanlari,ondalik_alanlari,karakter_alanlari,fiili_kullanim_amaclari 
from openpyxl.comments import Comment
import os

app = Flask(__name__)


import pandas as pd

def eksik_zorunlu_kontrolu(df: pd.DataFrame, zorunluluk_dict: dict, fiili_kullanim_amaclari: dict) -> pd.DataFrame:
    eksikler = []
    eksik_indexler = []

    for idx, row in df.iterrows():
        fiili_kod = row.get("Fiili Kullanım Amacı", None)

     

        if pd.isna(fiili_kod) or int(fiili_kod) not in fiili_kullanim_amaclari:
            continue

        fiili_kullanim = fiili_kullanim_amaclari[int(fiili_kod)]

        if fiili_kullanim not in zorunluluk_dict:
            continue

        
      

        zorunlu_alanlar = zorunluluk_dict[fiili_kullanim]
    
        for alan in df.columns:
            deger = row.get(alan, "")
       
            # Zorunlu alan kontrolü
            if alan in zorunlu_alanlar:
                if pd.isna(deger) or str(deger).strip() == "":
                    eksikler.append({
                        "Rapor No": row.get("Rapor No"),
                        "Fiili Kullanım": fiili_kullanim,
                        "Alan": alan,
                        "Hata": "Zorunlu Alan Boş",                      

                    })
                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Zorunlu Alan Boş"))
              # Tam sayı kontrolü (sadece boş olmayan alanlar için)
            if alan in tamsayi_alanlari:
                if not pd.isna(deger) and str(deger).strip() != "":
                    try:
                        sayi = float(str(deger).strip().replace(",", "."))
                        if not sayi.is_integer():
                            raise ValueError("Ondalık sayı girilmiş")

                         # Tam sayı kontrolü (sadece boş olmayan alanlar için) TCKN
                        if alan in ["Uzman (TCKN)", "Sorumlu Değerleme Uzmanı (TCKN)", "Denetmen (TCKN)"]:                              
                                try:
                                    
                                     if pd.isna(deger) or str(deger).strip() == "":
                                         raise ValueError("TCKN boş")
                                     if isinstance(deger, float):
                                         tckn_raw = str(int(deger))
                                     else:
                                         tckn_raw = str(deger).strip()
                                     if not tckn_raw.isdigit():
                                         raise ValueError("Sayısal değil")
                                     if len(tckn_raw) != 11:
                                         raise ValueError("11 haneli değil")
                                     if tckn_raw.startswith("0"):
                                          raise ValueError("0 ile başlamaz")                                   
                                except Exception:
                                    eksikler.append({                                                                 
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "11 Haneli Sayısal TCKN Bekleniyor"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"11 Haneli Sayısal TCKN Bekleniyor"))

                                 # Taşınmaz ID kontrolü (sadece boş olmayan alanlar için ve 8.Grup hariç)
                        if alan == "Taşınmaz ID":
                            try:
                                tasinmaz_id = int(float(str(deger).strip()))
                            except:
                                 tasinmaz_id = None
                           
                            if tasinmaz_id is None or not (1 <= tasinmaz_id <= 999999999):
                                eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                        "Hata": "1 ile 999999999 arasında geçerli bir Tam Sayı Taşınmaz ID bekleniyor"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"1 ile 999999999 arasında geçerli bir Tam Sayı Taşınmaz ID bekleniyor"))

                               # Ada ve Parsel kontrolü
                        if alan in ["Ada", "Parsel"]:
                            try:
                               sayisal_deger = int(float(str(deger).strip()))
                               if len(str(sayisal_deger)) > 50:
                                eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "50 karakteri aşan Ada/Parsel bilgisi"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"50 karakteri aşan Ada/Parsel bilgisi"))
                            except:
                                if sayisal_deger is None:
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı bir değer girilmeli (örn. 26 veya 1)"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı bir değer girilmeli (örn. 26 veya 1)"))
                        
                                 

                          # Arsa Payı ve Arsa Paydası kontrolü
                        if alan in ["Arsa Payı", "Arsa Paydası"]:
                          try: 
                               arsa_deger = int(float(str(deger).strip()))
                               if len(str(arsa_deger)) > 20:
                                 eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "20 karakteri aşan Arsa Payı/Paydası girilemez"
        
                                })  
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"20 karakteri aşan Arsa Payı/Paydası girilemez"))
                          except:
                               if arsa_deger == None:
                                eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı bir Arsa Payı/Paydası girilmelidir"
                                })
                               eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı bir Arsa Payı/Paydası girilmelidir"))
                       

                          # Yapı Kat Sayısı kontrolü (1-99 arası tam sayı, zorunlu)
                        if alan == "Yapı Kat Sayısı":
                          try:
                               kat_sayisi = int(float(str(deger).strip()))
                               if not (1 <= kat_sayisi <= 99):
                                   eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "Geçersiz değer: 1 ile 99 arasında bir sayı olmalı"
        
                                })
                               eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 1 ile 99 arasında bir sayı olmalı"))
                          except:
                                kat_sayisi = None
                                eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı bir değer girilmeli (1-99 arası)"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı bir değer girilmeli (1-99 arası)"))
                      
                          
                           # Salon, Banyo, Mutfak, Balkon kontrolü (0-9 arası adet, zorunlu)

                        if alan in ["Salon", "Banyo", "Mutfak", "Balkon"]:
                            try:
                               adet = int(float(str(deger).strip()))
                               if not (0 <= adet <= 9):
                                   eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "Geçersiz değer: 0 ile 9 arasında bir sayı olmalı"
        
                                })
                               eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 ile 9 arasında bir sayı olmalı"))
                            except:
                                adet = None
                                eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı değer girilmeli (0-9 adet aralığında)"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı değer girilmeli (0-9 adet aralığında)"))

                          # Salon, Banyo, Mutfak, Balkon kontrolü (0-9 arası adet, zorunlu)

                        if alan in ["Salon", "Banyo", "Mutfak", "Balkon"]:
                            try:
                               adet = int(float(str(deger).strip()))
                               if not (0 <= adet <= 9):
                                    eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "Geçersiz değer: 0 ile 9 arasında bir sayı olmalı"      
                                })
                               eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 ile 9 arasında bir sayı olmalı"))
                            except:
                                adet = None
                                eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı değer girilmeli (0-9 adet aralığında)"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı değer girilmeli (0-9 adet aralığında)"))

                         # Toplam Bağımsız Bölüm Sayısı kontrolü (zorunlu değil, ama varsa aralık kontrolü yapılır)

                        if alan == "Toplam Bağımsız Bölüm Sayısı":
                           if str(deger).strip() != "":
                            try:
                               bolum_sayisi = int(float(str(deger).strip()))
                            except:
                                bolum_sayisi = None
                                if bolum_sayisi is None or not (1 <= bolum_sayisi <= 999):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir sayı girilmeli (1-999 aralığında)"
                                }) 
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir sayı girilmeli (1-999 aralığında)"))



                           # Oda kontrolü (0-99 aralığında, zorunlu)

                        if alan == "Oda":
                            try:
                               oda_sayisi = int(float(str(deger).strip()))
                            except:
                                 oda_sayisi = None
                                 if not (0 <= oda_sayisi <= 99):
                                   eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "Geçersiz değer: 0 ile 99 arasında bir sayı olmalı"        
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 ile 99 arasında bir sayı olmalı"))
                            if oda_sayisi is None:
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı değer girilmeli (0-99 adet aralığında)"
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı değer girilmeli (0-99 adet aralığında)"))

                        # Dağıtımcı ile Kalan Anlaşma Süresi (Ay) kontrolü (zorunlu değil, ama varsa 0-999 arası olmalı)

                        if alan == "Dağıtımcı ile Kalan Anlaşma Süresi (Ay)":
                           if str(deger).strip() != "":
                            try:
                                sure = int(float(str(deger).strip()))
                                if not (0 <= oda_sayisi <= 99):
                                    eksikler.append({
                                   "Rapor No": row.get("Rapor No"),
                                   "Fiili Kullanım": fiili_kullanim,
                                   "Alan": alan,
                                   "Hata": "Geçerli bir sayı girilmeli (0-999 ay aralığında)"
                                })
                                eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı değer girilmeli ((0-999  adet aralığında)"))
                            except:
                                  sure = None

                            if sure is None or not (0 <= sure <= 999):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Tam Sayı değer girilmeli (0-99 adet aralığında)"
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı değer girilmeli (0-99 adet aralığında)"))

                        # Pompa Sayısı ve En Yakın İstasyona Mesafe (km) kontrolü

                        if alan in ["Pompa Sayısı", "En Yakın İstasyona Mesafe (km)"]:
                            try:
                                sayi = int(float(str(deger).strip()))
                            except:
                                  sayi = None

                            if sayi is None or not (0 <= sayi <= 99):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir değer girilmeli (0-99 aralığında)"
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir değer girilmeli (0-99 aralığında"))
                       # Mağaza Sayısı kontrolü

                        if alan == "Mağaza Sayısı":
                            try:
                                magaza_sayisi = int(float(str(deger).strip()))
                            except:
                                  magaza_sayisi = None

                            if magaza_sayisi is None or not (1 <= magaza_sayisi <= 9999):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir Mağaza Sayısı girilmeli (1-9999 adet aralığında)"
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir değer girilmeli (1-9999  aralığında"))
                          # Yıllık Ziyaretçi Sayısı kontrolü (zorunlu değil ama varsa 1-99.999.999 aralığında olmalı)

                        if alan == "Yıllık Ziyaretçi Sayısı":
                           if str(deger).strip() != "":
                            try:
                                ziyaretci = int(float(str(deger).replace(".", "").strip()))
                            except:
                                   ziyaretci = None

                            if ziyaretci is None or not (1 <= ziyaretci <= 99999999):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir sayı girilmeli (1 - 99.999.999 adet aralığında)"
                                })
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir sayı girilmeli (1 - 99.999.999 adet aralığında)"))

                          # Yıldız Sayısı kontrolü (zorunlu değil, ama varsa 0-7 aralığında olmalı)
                        if alan == "Yıldız Sayısı":
                           if str(deger).strip() != "":
                            try:
                                yildiz = int(float(str(deger).strip()))
                            except:
                                  yildiz = None
                            
                           if yildiz is None or not (0 <= yildiz <= 7):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir Yıldız Sayısı girilmeli (0-7 aralığında)"
                                }) 
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir Yıldız Sayısı girilmeli (0-7 aralığında)"))

                        # Oda Sayısı kontrolü (zorunlu, 1-9999 aralığında olmalı)
                        if alan == "Oda Sayısı":
                           if str(deger).strip() != "":
                            try:
                                oda_sayisi = int(float(str(deger).strip()))
                            except:
                                  oda_sayisi = None                        
                           if oda_sayisi is None or not (1 <= oda_sayisi <= 9999):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir Oda Sayısı girilmeli (1 - 9999 adet aralığında)"
                                }) 
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir Oda Sayısı girilmeli (1 - 9999 adet aralığında)"))
                        # Yıl İçinde Açık Olduğu Gün Sayısı
                        if alan == "Yıl İçinde Açık Olduğu Gün Sayısı":
                           if str(deger).strip() != "":
                            try:
                                gun_sayisi = int(float(str(deger).strip()))

                            except:
                                  gun_sayisi = None

                            
                           if gun_sayisi is None or not (0 <= gun_sayisi <= 366):
                                 eksikler.append({
                                    "Rapor No": row.get("Rapor No"),
                                    "Fiili Kullanım": fiili_kullanim,
                                    "Alan": alan,
                                    "Hata": "Geçerli bir gün sayısı girilmeli (0 - 366 aralığında)"
                                }) 
                                 eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçerli bir gün sayısı girilmeli (0 - 366 aralığında)"))
                         # Kalan Üst Hakkı Süresi (Ay)
                        if alan == "Kalan Üst Hakkı Süresi (Ay)":
                           if str(deger).strip() != "":
                               ust_hakki = row.get("Üst Hakkı Var mı?", "").strip().lower()
                           if ust_hakki == "1":
                                try:
                                    ust_sure = int(float(str(deger).strip()))
                                except:
                                      ust_sure = None

                            
                                if ust_sure is None or not (1 <= ust_sure <= 999):
                                     eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Üst Hakkı Var seçildiğinde 1-999 ay aralığında bir süre girilmeli"
                                    })
                                     eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Üst Hakkı Var seçildiğinde 1-999 ay aralığında bir süre girilmeli"))                                               
                    except:
                        eksikler.append({
                            "Rapor No": row.get("Rapor No"),
                            "Fiili Kullanım": fiili_kullanim,
                            "Alan": alan,
                            "Hata": "Tam Sayı Bekleniyor"
                        })
                        eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı Bekleniyor"))
            # Ondalık sayı kontrolü (sadece boş olmayan alanlar için ve sarı alanlar hariç)
            if alan in ondalik_alanlari:
                    if not pd.isna(deger) and str(deger).strip() != "":
                        try:
                            sayi = float(str(deger).strip().replace(",", "").replace("₺", "").replace("$", ""))

                        

                             # Ortak gruptakiler için temel aralık kontrolü
                            if alan in ["Parsel Yüzölçümü (m²)",  "Enlem", "Boylam",  "Yasal Brüt Kullanım Alanı (m²)","Mevcut Brüt Kullanım Alanı (m²)"]:
                            
                                if alan == "Enlem" and not (35.0000 <= sayi <= 43.0000):
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Geçersiz değer: 35.0000 - 43.0000 aralığında olmalı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 35.0000 - 43.0000 aralığında olmalı"))
                             
                                elif alan == "Boylam" and not (25.0000 <= sayi <= 45.0000):
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Geçersiz değer: 25.0000 - 45.0000 aralığında olmalı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 25.0000 - 45.0000  aralığında olmalı"))
                               
                                elif alan not in ["Enlem", "Boylam"] and not (0 <= sayi <= 999999999.99):
                                     eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Geçersiz değer: 0 - 999.999.999,99 aralığında olmalı"
                                    })
                                     eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 - 999.999.999,99  aralığında olmalı"))
                                elif alan == "Kiralanabilir Alan (m²)":
                                    if not (0 <= sayi <= 999999999.99):
                                      eksikler.append({
                                       "Rapor No": row.get("Rapor No"),
                                       "Fiili Kullanım": fiili_kullanim,
                                       "Alan": alan,
                                      "Hata": "Geçersiz değer: 0 - 999.999.999,99 m² aralığında olmalı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 - 999.999.999,99 m² aralığında olmalı"))

                              # Arsa, Arazi grubu için
                            elif alan == "Terk Sonrası Parselin Yüz Ölçümü (m²)":
                                if not (0 <= sayi <= 999999999.99):
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Geçersiz değer: 0 - 999.999.999,99 m² aralığında olmalı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 - 999.999.999,99 m² aralığında olmalı"))

                            # Arsa ve Arazi Hariç grubu için
                            elif alan == "Arsa Birim Değeri (TL/m²)":
                                if not (0 <= sayi <= 999999999.99):
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Geçersiz değer: 0 - 999.999.999,99 aralığında olmalı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Geçersiz değer: 0 - 999.999.999,99 aralığında olmalı"))
                                                                                
                        except:
                            eksikler.append({
                                "Rapor No": row.get("Rapor No"),
                                "Fiili Kullanım": fiili_kullanim,
                                "Alan": alan,
                                "Hata": "Tam Sayı Bekleniyor"
                            })
                            eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Tam Sayı Bekleniyor"))
            # Karakter bazlı alan kontrolü
            if alan in karakter_alanlari :
                    if not pd.isna(deger) and str(deger).strip() != "":
                      try:
                            temiz_deger = str(deger).strip().replace(",", "").replace("₺", "").replace("$", "")
                      except:


                          #Mahalle/Köy (Takbis)", "Mevcut Mahalle/Köy
                            if alan in ["Mahalle/Köy (Takbis)", "Mevcut Mahalle/Köy"]:
                                try:
                                    deger_str = str(deger).strip().replace(",", "").replace("₺", "").replace("$", "")
                                except:
                                 deger_str = str(deger).strip().replace(",", "").replace("₺", "").replace("$", "") 
                                 if len(deger_str) > 50:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": f"{alan} alanı için maksimum 50 karakter sınırı aşıldı"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"alanı için maksimum 50 karakter sınırı aşıldı"))

                               
                                     #Ana Taşınmaz Niteliği
                            if alan == "Ana Taşınmaz Niteliği":
                                try:
                                    deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if not deger_str or len(deger_str) > 250:
                                   if tasinmaz_id is None or not (1 <= tasinmaz_id <= 999999999):
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                        "Hata": "Boş olamaz ve en fazla 250 karakter/sayı içerebilir"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"Boş olamaz ve en fazla 250 karakter/sayı içerebilir"))
                             
                               #BLOK                       
                            if alan == "Blok":
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 20:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 20 karakter içerebilir"
                                    })
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 20 karakter içerebilir"))
                              #KAT
                            elif alan == "Kat":
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 20:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 20 karakter içerebilir"
                                    }) 
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 20 karakter içerebilir"))

                              #BAĞIMSIZ BÖLÜM NO
                            elif alan == "Bağımsız Bölüm Numarası":
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 20:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 20 karakter içerebilir"
                                    }) 
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 20 karakter içerebilir"))

                             # 50 karakter sınırı için örnek kontrol (Bulvar/Cadde ve Sokak)
                            if alan in ["Bulvar/Cadde", "Sokak"]:
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 50:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 50 karakter içerebilir"
                                    })   
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 50 karakter içerebilir"))

                             # 250 karakter sınırı için kontrol (Dağıtımcı Markası)
                            elif alan == "Dağıtımcı Markası":
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 250:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 250 karakter içerebilir"
                                    })   
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 250 karakter içerebilir"))
                      
                             # 250 karakter sınırı için kontrol (Bağımsız Bölüm Niteliği)
                            elif alan == "Bağımsız Bölüm Niteliği":
                                try:
                                   deger_str = str(deger).strip()
                                except:
                                     deger_str = ""
                                if len(deger_str) > 250:
                                    eksikler.append({
                                        "Rapor No": row.get("Rapor No"),
                                        "Fiili Kullanım": fiili_kullanim,
                                        "Alan": alan,
                                         "Hata": "En fazla 250 karakter içerebilir"
                                    }) 
                                    eksik_indexler.append((idx + 2, df.columns.get_loc(alan) + 1,"En fazla 250 karakter içerebilir"))

    return pd.DataFrame(eksikler), eksik_indexler

   
def boya_ve_kaydet(file_stream, eksik_indexler, cikti_yolu="static/boyali.xlsx"):
    sari = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Excel'i hem pandas hem openpyxl için oku
    file_stream.seek(0)
    df = pd.read_excel(file_stream)
    file_stream.seek(0)
    wb = load_workbook(file_stream)
    ws = wb.active

    # Eksik hücreleri doğrudan satır/sütun indexine göre boya
    for satir_index, sutun_index,yorum in eksik_indexler:
        excel_satir = satir_index  # zaten 1-based geliyorsa bu doğru
        excel_sutun = sutun_index

        hucre = ws.cell(row=excel_satir, column=excel_sutun)
        hucre.fill = sari
        hucre.comment = Comment(yorum, "Sistem")

    # Belleğe yaz
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Diske yaz
    with open(cikti_yolu, "wb") as f:
        f.write(output.read())

    return cikti_yolu


@app.route('/', methods=['GET', 'POST'])
def index():
    results = []
    eksik_alanlar_df = pd.DataFrame()

    if request.method == 'POST':
        file = request.files['gabim_file']
        if file and file.filename.endswith('.xlsx'):
            # file.stream yerine BytesIO kopyası al
            file_stream = BytesIO(file.read())
            file_stream.seek(0)

            # DataFrame'i oku
            df = pd.read_excel(file_stream)

            # Zorunlu alan kontrolü
            eksik_alanlar_df, eksik_indexler = eksik_zorunlu_kontrolu(df, zorunluluk_dict, fiili_kullanim_amaclari)

            # 1. Eksik alan listesini Excel'e yaz
            output = BytesIO()
            eksik_alanlar_df.to_excel(output, index=False)
            output.seek(0)
            with open("static/sonuc.xlsx", "wb") as f:
                f.write(output.read())

            # 2. Orijinal dosyayı eksik yerlere sarı boya ile kaydet
            file_stream.seek(0)
            boya_ve_kaydet(file_stream, eksik_indexler)

            # HTML'de göstermek için sözlük formatına çevir
            results = eksik_alanlar_df.to_dict(orient='records')
            return render_template("index.html", results=results)

        else:
            return "Yalnızca .xlsx uzantılı dosyalar destekleniyor.", 400

    return render_template("index.html", results=results)

@app.route('/excel_indir')
def excel_indir():
    return send_file("static/sonuc.xlsx", as_attachment=True)
